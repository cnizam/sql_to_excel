import pandas as pd
from connectOracleDB import conn
import openpyxl as xl
import os
import datetime
import warnings
import time
import shutil

warnings.filterwarnings("ignore", category=DeprecationWarning)
lighblue_color = 'dce6f1'  #you can change the row coloring here
lightblue_fill = xl.styles.PatternFill(start_color=lighblue_color, end_color=lighblue_color, fill_type='solid')

today_date = datetime.datetime.today().strftime('%F')

#For the code to work, the file path must be changed to suit your own use
sql_path = r'\put_your_sql_file_here'  # folder where sql query files are put
excel_path = r'\excel_path'            # folder where excel file was created
sql_archive_path = r'\sql_archive'     # folder where it is archived after the query is run

file_list = os.listdir(sql_path)

thin_border = xl.styles.borders.Border(left=xl.styles.borders.Side(style='thin'), 
                     right=xl.styles.borders.Side(style='thin'), 
                     top=xl.styles.borders.Side(style='thin'), 
                     bottom=xl.styles.borders.Side(style='thin'))


files_only = [file for file in file_list if os.path.isfile(os.path.join(sql_path, file))]

for file in files_only:
    file_name = os.path.splitext(file)[0]
    os.chdir(sql_path)
    fd = open(file, 'r')
    file = fd.read()
    fd.close()
    data = pd.read_sql(file, conn, index_col=None)
    os.chdir(excel_path)
    data.to_excel(f'{file_name} - {today_date}.xlsx', index = False)
    time.sleep(3)
    wb = xl.load_workbook(filename = f'{file_name} - {today_date}.xlsx', read_only = False)
    ws = wb.active
    max_column = wb.active.max_column
    max_row = wb.active.max_row
    date_format = '%d.%m.%Y'

    for col in range(0,max_column):
        if  data.iloc[:,col].dtype == 'int64' and 'ID' not in data.iloc[:,col].name and 'id' not in data.iloc[:,col].name and 'unique' not in data.iloc[:,col].name and 'identity' not in data.iloc[:,col].name:
            for i in range(2,max_row+1):
                _cell = ws.cell(row = i, column=col+1)
                _cell.number_format = '#,##0'
                _cell.alignment = xl.styles.Alignment(wrap_text=True,vertical='center', horizontal='center') 
        elif data.iloc[:,col].dtype == 'float64':
            if 'RATE' in data.iloc[:,col].name or 'rate' in data.iloc[:,col].name or 'Rate' in data.iloc[:,col].name or 'percent' in data.iloc[:,col].name or 'Percent' in data.iloc[:,col].name:
                for i in range(2,max_row+1):
                    _cell = ws.cell(row = i, column=col+1)
                    _cell.number_format = '0.00%'
                    _cell.alignment = xl.styles.Alignment(wrap_text=True,vertical='center', horizontal='center')
            else:
                for i in range(2,max_row+1):
                    _cell = ws.cell(row = i, column=col+1)
                    _cell.number_format = '#,##0'
                    _cell.alignment = xl.styles.Alignment(wrap_text=True,vertical='center', horizontal='right')
        elif data.iloc[:,col].dtype == 'datetime64[ns]':
            for i in range(2,max_row+1):
                _cell = ws.cell(row = i, column=col+1)
                _cell.value = _cell.value.strftime(date_format)
                _cell.alignment = xl.styles.Alignment(wrap_text=True,vertical='center', horizontal='center')    
        elif data.iloc[:,col].dtype == 'object':
            for i in range(2,max_row+1):
                _cell = ws.cell(row = i, column=col+1)
                _cell.alignment = xl.styles.Alignment(wrap_text=True,vertical='center', horizontal='left')
        else:
            for i in range(2,max_row+1):
                _cell = ws.cell(row = i, column=col+1)
                _cell.alignment = xl.styles.Alignment(wrap_text=True,vertical='center', horizontal='center')
    for i in range(1,max_column+1):
        ws.cell(row=1, column=i).fill = xl.styles.PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type = "solid")
        ws.cell(row=1, column=i).font = xl.styles.Font(bold=True,color='ffffff')
        ws.cell(row=1, column=i).alignment = xl.styles.Alignment(wrap_text=True,vertical='center', horizontal='center')
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value
        if ws.column_dimensions[col].width < 15:
            ws.column_dimensions[col].width = 15

    formula_rule = xl.formatting.rule.FormulaRule(formula=['MOD(ROW(),2)'], stopIfTrue=False, fill=lightblue_fill)
    column_letter = xl.utils.get_column_letter(max_column)
    cell_range = f'A2:{column_letter}{max_row}'
    ws.conditional_formatting.add(cell_range, formula_rule)

    for i in range(1,max_row+1):
        for j in range(1,max_column+1):
            ws.cell(row=i, column=j).border = thin_border

    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    wb.save(f'{file_name} - {today_date}.xlsx')

    source_file = sql_path+f"\{file_name}.sql"
    destination_file = sql_archive_path+f"\{file_name} - {today_date}.sql"

    try:
        shutil.copyfile(source_file, destination_file)
        print(f"File copied successfully from {source_file} to {destination_file}")
    except FileNotFoundError:
        print("Source file not found.")
    except PermissionError:
        print("Permission denied. Check if you have write access to the destination directory.")
    except Exception as e:
        print(f"An error occurred: {e}")
   
    try:
        if os.path.exists(source_file):

            os.remove(source_file)
            print(f"File {source_file} deleted successfully.")
        else:
            print(f"File {source_file} not found.")
    except Exception as e:
        print(f"An error occurred: {e}")
